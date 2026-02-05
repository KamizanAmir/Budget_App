import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import datetime
import os

# --- CONFIGURATION ---
FILE_NAME = "my_budget_data.xlsx"
st.set_page_config(page_title="My Personal Budget", page_icon="💰", layout="wide")

# --- HELPER FUNCTIONS ---
def load_data():
    """Loads all sheets from the Excel file."""
    try:
        return pd.read_excel(FILE_NAME, sheet_name=None)
    except FileNotFoundError:
        return None

def save_row(sheet_name, data):
    """Appends a list of values to the specified Excel sheet."""
    wb = load_workbook(FILE_NAME)
    ws = wb[sheet_name]
    ws.append(data)
    wb.save(FILE_NAME)

def delete_row(sheet_name, row_index):
    """Deletes a row from Excel by index (row_index is 0-based index from DataFrame)."""
    wb = load_workbook(FILE_NAME)
    ws = wb[sheet_name]
    # Excel rows are 1-based. Row 1 is header. Data starts at Row 2.
    # So DataFrame index 0 corresponds to Excel Row 2.
    # We delete row_index + 2.
    ws.delete_rows(row_index + 2)
    wb.save(FILE_NAME)

# --- SUCCESS MESSAGE HANDLING ---
# This ensures the message survives the page reload (st.rerun)
if 'success_msg' in st.session_state:
    st.success(st.session_state['success_msg'])
    del st.session_state['success_msg']

st.title("💰 My Personal Budget Tracker")

# Load current data
all_data = load_data()

if all_data is None:
    st.error(f"Error: {FILE_NAME} not found. Please run setup_storage.py first.")
    st.stop()

# --- TABS ---
# Reordered: Income First
tab1, tab2, tab3 = st.tabs(["📥 Add Income", "💸 Add Expense", "📊 Analytics & History"])

# --- TAB 1: ADD INCOME ---
with tab1:
    st.header("New Income Entry")
    with st.form("income_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            date_inc = st.date_input("Date", datetime.date.today(), key="inc_date")
            source_inc = st.text_input("Source (e.g., Salary, Freelance)")
        with col2:
            amount_inc = st.number_input("Amount (RM)", min_value=0.0, format="%.2f", key="inc_amt")
            
        submitted_inc = st.form_submit_button("Save Income", type="primary")
        
        if submitted_inc:
            save_row('Income', [date_inc, source_inc, amount_inc])
            st.session_state['success_msg'] = f"✅ Income Saved: RM{amount_inc} from {source_inc}"
            st.rerun()

# --- TAB 2: ADD EXPENSE ---
with tab2:
    st.header("New Expense Entry")
    with st.form("expense_form", clear_on_submit=True):
        col1, col2 = st.columns(2)
        with col1:
            date_exp = st.date_input("Date", datetime.date.today())
            category_exp = st.selectbox("Category", ["Food", "Transport", "Utilities", "Shopping", "Housing", "Other"])
        with col2:
            amount_exp = st.number_input("Amount (RM)", min_value=0.0, format="%.2f")
            desc_exp = st.text_input("Description (e.g., Nasi Lemak)")
        
        submitted_exp = st.form_submit_button("Save Expense", type="primary")
        
        if submitted_exp:
            save_row('Expenses', [date_exp, desc_exp, category_exp, amount_exp])
            st.session_state['success_msg'] = f"✅ Expense Saved: {desc_exp} (RM{amount_exp})"
            st.rerun()

# --- TAB 3: ANALYTICS & HISTORY (View, Filter, Delete) ---
with tab3:
    st.header("Monthly Overview")

    # 1. Prepare Data
    df_inc = all_data['Income']
    df_exp = all_data['Expenses']
    
    # Convert Date columns to datetime for filtering
    df_inc['Date'] = pd.to_datetime(df_inc['Date'])
    df_exp['Date'] = pd.to_datetime(df_exp['Date'])

    # 2. Month Filter
    # Get unique months from both datasets
    all_dates = pd.concat([df_inc['Date'], df_exp['Date']]).dropna()
    
    if not all_dates.empty:
        # --- FIX: Use drop_duplicates() instead of unique() so we can sort ---
        month_years = all_dates.dt.to_period('M').drop_duplicates().sort_values(ascending=False)
        
        selected_period = st.selectbox("Select Month", month_years)
        
        # Filter DataFrames based on selection
        mask_inc = df_inc['Date'].dt.to_period('M') == selected_period
        mask_exp = df_exp['Date'].dt.to_period('M') == selected_period
        
        filtered_inc = df_inc[mask_inc].copy()
        filtered_exp = df_exp[mask_exp].copy()
        
        # 3. Calculate Totals
        total_income = filtered_inc['Amount'].sum()
        total_expense = filtered_exp['Amount'].sum()
        balance = total_income - total_expense

        # 4. Display Metrics
        m1, m2, m3 = st.columns(3)
        m1.metric("Total Income", f"RM {total_income:,.2f}")
        m2.metric("Total Expenses", f"RM {total_expense:,.2f}")
        m3.metric("Remaining Balance", f"RM {balance:,.2f}", delta_color="normal")
        
        st.divider()

        # 5. Detailed View & Delete Functionality
        c1, c2 = st.columns(2)
        
        with c1:
            st.subheader("Expenses Details")
            if not filtered_exp.empty:
                for idx, row in filtered_exp.iterrows():
                    col_text, col_btn = st.columns([4, 1])
                    with col_text:
                        st.text(f"{row['Date'].date()} | {row['Category']} | {row['Description']} | RM{row['Amount']}")
                    with col_btn:
                        if st.button("🗑", key=f"del_exp_{idx}", help="Delete this entry"):
                            delete_row('Expenses', idx)
                            st.session_state['success_msg'] = "❌ Entry Deleted!"
                            st.rerun()
                
                st.download_button(
                    label="Download Expenses (CSV)",
                    data=filtered_exp.to_csv(index=False),
                    file_name=f"expenses_{selected_period}.csv",
                    mime="text/csv"
                )
            else:
                st.info("No expenses found for this month.")

        with c2:
            st.subheader("Income Details")
            if not filtered_inc.empty:
                for idx, row in filtered_inc.iterrows():
                    col_text, col_btn = st.columns([4, 1])
                    with col_text:
                        st.text(f"{row['Date'].date()} | {row['Source']} | RM{row['Amount']}")
                    with col_btn:
                        if st.button("🗑", key=f"del_inc_{idx}", help="Delete this entry"):
                            delete_row('Income', idx)
                            st.session_state['success_msg'] = "❌ Entry Deleted!"
                            st.rerun()
                            
                st.download_button(
                    label="Download Income (CSV)",
                    data=filtered_inc.to_csv(index=False),
                    file_name=f"income_{selected_period}.csv",
                    mime="text/csv"
                )
            else:
                st.info("No income found for this month.")

    else:
        st.info("No data available yet. Add some entries!")